import openpyxl
import json
import os

excel_path = "DUMMY_MASTER DATA DESA PESISIR.xlsx"
json_output_dir = os.path.join("src", "data")
json_output_path = os.path.join(json_output_dir, "desa_pesisir.json")

# Create output dir if not exists
os.makedirs(json_output_dir, exist_ok=True)

print(f"Loading Excel workbook: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)

data = {
    "wilayah": {},
    "demografi": [],
    "sosial_ekonomi": [],
    "data_kapal": [],
    "data_alat_tangkap": [],
    "data_kelompok": [],
    "data_infrastruktur": []
}

# 1. Parse MASTER WILAYAH
ws_wil = wb['MASTER WILAYAH']
for r in list(ws_wil.iter_rows(values_only=True))[1:]:
    if not r[1]: # Check ID_DESA
        continue
    data["wilayah"][r[1]] = {
        "desa": r[0],
        "id_desa": r[1],
        "latitude": r[2],
        "longitude": r[3],
        "kecamatan": r[4],
        "kabupaten": r[5],
        "provinsi": r[6],
        "keterangan_kk": r[7],
        "nama_kk": r[8]
    }

# 2. Parse DEMOGRAFI
ws_demo = wb['DEMOGRAFI']
for r in list(ws_demo.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["demografi"].append({
        "id_desa": r[0],
        "desa": r[1],
        "tahun": r[2],
        "laki_laki": r[3],
        "perempuan": r[4],
        "total_penduduk": r[5],
        "jumlah_kk": r[6],
        "jumlah_dusun": r[7],
        "luas_wilayah": r[8],
        "sumber": r[9]
    })

# 3. Parse SOSIAL EKONOMI
ws_se = wb['SOSIAL EKONOMI']
for r in list(ws_se.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["sosial_ekonomi"].append({
        "id_desa": r[0],
        "desa": r[1],
        "umkm": r[2] or 0,
        "nelayan": r[3] or 0,
        "pembudidaya": r[4] or 0,
        "pengolah_pemasar": r[5] or 0,
        "kapal": r[6] or 0,
        "alat_tangkap": r[7] or 0,
        "kelompok": r[8] or 0,
        "bumdes": r[9],
        "koperasi": r[10],
        "upi": r[11],
        "wisata_alam": r[12],
        "wisatawan": r[13],
        "masyarakat_adat": r[14],
        "kearifan_lokal": r[15],
        "infrastruktur": r[16] or 0,
        "sumber": r[17]
    })

# 4. Parse DATA_KAPAL
ws_kapal = wb['DATA_KAPAL']
for r in list(ws_kapal.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["data_kapal"].append({
        "id_desa": r[0],
        "desa": r[1],
        "tahun": r[2],
        "jenis": r[3],
        "jumlah": r[4] or 0,
        "sumber": r[5]
    })

# 5. Parse DATA_ALAT_TANGKAP
ws_alat = wb['DATA_ALAT_TANGKAP']
for r in list(ws_alat.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["data_alat_tangkap"].append({
        "id_desa": r[0],
        "desa": r[1],
        "tahun": r[2],
        "jenis": r[3],
        "jumlah": r[4] or 0,
        "sumber": r[5]
    })

# 6. Parse DATA_KELOMPOK
ws_kel = wb['DATA_KELOMPOK']
for r in list(ws_kel.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["data_kelompok"].append({
        "id_desa": r[0],
        "desa": r[1],
        "jenis": r[2],
        "nama": r[3],
        "tahun_berdiri": r[4],
        "status": r[5],
        "sumber": r[6]
    })

# 7. Parse DATA_INFRASTRUKTUR_PERIKANAN
ws_infra = wb['DATA_INFRASTRUKTUR_PERIKANAN']
for r in list(ws_infra.iter_rows(values_only=True))[1:]:
    if not r[0]: # Check ID_DESA
        continue
    data["data_infrastruktur"].append({
        "id_desa": r[0],
        "desa": r[1],
        "tahun": r[2],
        "jenis": r[3],
        "jumlah": r[4] or 0,
        "sumber": r[5]
    })

print(f"Writing parsed JSON data to: {json_output_path}")
with open(json_output_path, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("Conversion complete!")
